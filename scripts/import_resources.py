from __future__ import annotations

import json
import re
from datetime import datetime, date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RESOURCES = ROOT / "resources"
STORE = ROOT / "data" / "store.json"


def default_users() -> list[dict[str, str]]:
    return [
        {"id": "u-admin", "username": "admin", "password": "admin123", "name": "Administrator", "role": "admin"},
        {"id": "u-requester", "username": "requester", "password": "request123", "name": "Requisition User", "role": "requester"},
        {"id": "u-store", "username": "store", "password": "store123", "name": "Store / PMU Officer", "role": "store"},
        {"id": "u-approver", "username": "approver", "password": "approve123", "name": "Final Approver", "role": "approver"},
    ]


def clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def as_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean(value)
    return text


def as_num(value: Any) -> float:
    try:
        if value is None or value == "":
            return 0
        return float(value)
    except Exception:
        return 0


def slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")[:48] or "item"


def load_store() -> dict[str, Any]:
    if STORE.exists():
        return json.loads(STORE.read_text(encoding="utf-8"))
    return {
        "meta": {"name": "Inventory System", "createdAt": datetime.utcnow().isoformat()},
        "users": default_users(),
        "projects": [],
        "budgetHeads": [],
        "infrastructures": [],
        "items": [],
        "requisitions": [],
        "receipts": [],
        "issues": [],
        "ledger": [],
        "expenses": [],
        "counters": {"requisition": 1, "receipt": 1, "issue": 1, "movement": 1, "expense": 1},
    }


def counter(store: dict[str, Any], key: str, prefix: str) -> str:
    store.setdefault("counters", {}).setdefault(key, 1)
    value = f"{prefix}-{store['counters'][key]:05d}"
    store["counters"][key] += 1
    return value


def ensure_default_users(store: dict[str, Any]):
    store.setdefault("users", [])
    existing = {clean(user.get("username", "")).lower() for user in store["users"]}
    for user in default_users():
        if user["username"] not in existing:
            store["users"].append(user)
            existing.add(user["username"])


def get_project(store: dict[str, Any], name: str, budget: float = 0) -> str:
    name = clean(name)
    if not name:
        name = "Unassigned"
    for project in store["projects"]:
        if project["name"].lower() == name.lower():
            if budget and not project.get("budget"):
                project["budget"] = budget
            return project["id"]
    project = {"id": f"p-{slug(name)}", "name": name, "budget": budget}
    store["projects"].append(project)
    return project["id"]


def get_item(store: dict[str, Any], name: str, category: str = "", unit: str = "") -> dict[str, Any]:
    name = clean(name)
    category = clean(category) or "General"
    unit = clean(unit) or "Nos"
    if not name:
        name = "Unknown Item"
    for item in store["items"]:
        if item["name"].lower() == name.lower() and item.get("unit", "").lower() == unit.lower():
            if category and item.get("category") in ("", "General"):
                item["category"] = category
            return item
    item = {"id": f"i-{slug(name)}-{len(store['items']) + 1}", "name": name, "category": category, "unit": unit}
    store["items"].append(item)
    return item


def add_movement(store: dict[str, Any], *, movement_type: str, date_value: str, item: dict[str, Any], quantity: float, unit: str, project_id: str, reference_type: str, reference_id: str, document_no: str, remarks: str):
    store["ledger"].append(
        {
            "id": counter(store, "movement", "MOV"),
            "date": date_value,
            "type": movement_type,
            "itemId": item["id"],
            "itemName": item["name"],
            "category": item.get("category", "General"),
            "quantity": quantity,
            "unit": unit or item.get("unit", ""),
            "projectId": project_id,
            "referenceType": reference_type,
            "referenceId": reference_id,
            "documentNo": document_no,
            "remarks": remarks,
        }
    )


def reset_imported(store: dict[str, Any]):
    store["projects"] = []
    store["budgetHeads"] = []
    store["infrastructures"] = []
    store["items"] = []
    store["requisitions"] = []
    store["receipts"] = []
    store["issues"] = []
    store["ledger"] = []
    store["expenses"] = []
    store["counters"] = {"requisition": 1, "receipt": 1, "issue": 1, "movement": 1, "expense": 1}


def import_oap(store: dict[str, Any]):
    wb = load_workbook(RESOURCES / "Yarju OAP Inventory Record.xlsx", data_only=True)
    expense_sheets = [
        "PMU Road Construction",
        "PMU Temp.sched",
        "PMU Construction 15m",
        "PMU Construction 20+18M",
        "PMU Construction 50M",
        "Mushroom Shed",
    ]
    for sheet_name in expense_sheets:
        ws = wb[sheet_name]
        budget = as_num(ws["H3"].value)
        project_id = get_project(store, sheet_name, budget)
        for row in ws.iter_rows(min_row=5, values_only=True):
            item_name = clean(row[7] if len(row) > 7 else "")
            if not item_name:
                continue
            item = get_item(store, item_name, clean(row[8] if len(row) > 8 else ""), clean(row[10] if len(row) > 10 else ""))
            qty = as_num(row[9] if len(row) > 9 else 0)
            rate = as_num(row[11] if len(row) > 11 else 0)
            amount = as_num(row[12] if len(row) > 12 else 0)
            if not qty and not amount:
                continue
            store["expenses"].append(
                {
                    "id": counter(store, "expense", "EXP"),
                    "source": sheet_name,
                    "projectId": project_id,
                    "billNo": clean(row[1] if len(row) > 1 else ""),
                    "billDate": as_date(row[2] if len(row) > 2 else ""),
                    "dvNo": clean(row[3] if len(row) > 3 else ""),
                    "dvDate": as_date(row[4] if len(row) > 4 else ""),
                    "challanNo": clean(row[5] if len(row) > 5 else ""),
                    "challanDate": as_date(row[6] if len(row) > 6 else ""),
                    "itemId": item["id"],
                    "itemName": item["name"],
                    "category": item.get("category", ""),
                    "quantity": qty,
                    "unit": item.get("unit", ""),
                    "rate": rate,
                    "amount": amount,
                    "enterprise": clean(row[13] if len(row) > 13 else ""),
                    "purpose": clean(row[14] if len(row) > 14 else ""),
                }
            )

    for sheet_name in ["PIU Fuel & Lube Received", "PIU Stock Received"]:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            item_name = clean(row[3] if len(row) > 3 else "")
            qty = as_num(row[4] if len(row) > 4 else 0)
            if not item_name or not qty:
                continue
            item = get_item(store, item_name, "Fuel & Lube" if "Fuel" in sheet_name else "Store Stock", clean(row[5] if len(row) > 5 else ""))
            receipt = {
                "id": counter(store, "receipt", "REC"),
                "date": as_date(row[0] if len(row) > 0 else ""),
                "requisitionId": "",
                "projectId": "",
                "supplier": clean(row[1] if len(row) > 1 else ""),
                "challanNo": clean(row[2] if len(row) > 2 else ""),
                "challanDate": "",
                "dvNo": "",
                "dvDate": "",
                "billNo": "",
                "billDate": "",
                "receivedBy": "import",
                "receivedByName": "Imported Excel record",
                "remarks": clean(row[6] if len(row) > 6 else ""),
                "lines": [{"itemId": item["id"], "itemName": item["name"], "specification": "", "quantity": qty, "unit": item["unit"], "rate": 0, "amount": 0, "remarks": ""}],
            }
            store["receipts"].append(receipt)
            add_movement(store, movement_type="RECEIPT", date_value=receipt["date"], item=item, quantity=qty, unit=item["unit"], project_id="", reference_type="receipt", reference_id=receipt["id"], document_no=receipt["challanNo"], remarks=receipt["remarks"])

    ws = wb["Requestion Follow Up"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        item_name = clean(row[4] if len(row) > 4 else "")
        qty = as_num(row[6] if len(row) > 6 else 0)
        if not item_name or not qty:
            continue
        item = get_item(store, item_name, "Requisition Item", clean(row[7] if len(row) > 7 else ""))
        req_id = counter(store, "requisition", "REQ")
        store["requisitions"].append(
            {
                "id": req_id,
                "requisitionNo": clean(row[2] if len(row) > 2 else req_id),
                "requestDate": as_date(row[1] if len(row) > 1 else ""),
                "receivedDate": "",
                "projectId": "",
                "purpose": clean(row[11] if len(row) > 11 else ""),
                "status": "IMPORTED_FOLLOW_UP",
                "createdBy": "import",
                "createdByName": "Imported Excel record",
                "createdAt": datetime.utcnow().isoformat(),
                "supplyOrderNo": clean(row[3] if len(row) > 3 else ""),
                "approvals": [],
                "lines": [
                    {
                        "id": f"rl-{req_id}",
                        "itemId": item["id"],
                        "itemName": item["name"],
                        "specification": clean(row[5] if len(row) > 5 else ""),
                        "quantity": qty,
                        "unit": item["unit"],
                        "issuedTillDate": 0,
                        "balance": as_num(row[9] if len(row) > 9 else 0),
                        "remarks": clean(row[8] if len(row) > 8 else ""),
                    }
                ],
            }
        )


def import_issue_sheet(store: dict[str, Any], sheet_name: str, config: dict[str, int | str]):
    wb = load_workbook(RESOURCES / "Yarju Stone, Cement and others record.xlsx", data_only=True)
    ws = wb[sheet_name]
    start_row = int(config.get("start_row", 3))
    project_id = get_project(store, "Yarju OAP Site Consumption", 0)
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        item_name = clean(row[int(config["item_col"])] if len(row) > int(config["item_col"]) else "")
        qty = as_num(row[int(config["qty_col"])] if len(row) > int(config["qty_col"]) else 0)
        if not item_name or not qty:
            continue
        unit = clean(row[int(config["unit_col"])] if len(row) > int(config["unit_col"]) else config.get("unit", "Nos"))
        item = get_item(store, item_name, "Issued Construction Material", unit)
        issue = {
            "id": counter(store, "issue", "ISS"),
            "date": as_date(row[int(config["date_col"])] if len(row) > int(config["date_col"]) else ""),
            "projectId": project_id,
            "issueChallanNo": clean(row[int(config["issue_col"])] if "issue_col" in config and len(row) > int(config["issue_col"]) else ""),
            "issuedTo": clean(row[int(config["to_col"])] if "to_col" in config and len(row) > int(config["to_col"]) else ""),
            "issuedBy": "import",
            "issuedByName": clean(row[int(config["by_col"])] if "by_col" in config and len(row) > int(config["by_col"]) else "Imported Excel record"),
            "remarks": clean(row[int(config["remarks_col"])] if "remarks_col" in config and len(row) > int(config["remarks_col"]) else ""),
            "lines": [{"itemId": item["id"], "itemName": item["name"], "specification": clean(row[int(config["spec_col"])] if "spec_col" in config and len(row) > int(config["spec_col"]) else ""), "quantity": qty, "unit": unit, "remarks": ""}],
        }
        store["issues"].append(issue)
        add_movement(store, movement_type="ISSUE", date_value=issue["date"], item=item, quantity=qty, unit=unit, project_id=project_id, reference_type="issue", reference_id=issue["id"], document_no=issue["issueChallanNo"], remarks=issue["remarks"])


def import_issues(store: dict[str, Any]):
    configs = {
        "Boulders": {"start_row": 3, "date_col": 1, "issue_col": 5, "item_col": 6, "qty_col": 7, "unit_col": 8, "by_col": 9, "to_col": 10, "remarks_col": 11},
        "Sand": {"start_row": 3, "date_col": 1, "issue_col": 5, "item_col": 6, "qty_col": 7, "unit_col": 8, "to_col": 10, "remarks_col": 11},
        "Timber": {"start_row": 2, "date_col": 1, "issue_col": 5, "item_col": 6, "qty_col": 7, "unit_col": 8, "by_col": 9, "to_col": 10, "remarks_col": 11},
        "Aggregates": {"start_row": 3, "date_col": 1, "issue_col": 5, "item_col": 6, "qty_col": 9, "unit_col": 8, "to_col": 10, "remarks_col": 11},
        "Cement": {"start_row": 3, "date_col": 1, "issue_col": 5, "item_col": 6, "qty_col": 7, "unit_col": 8, "by_col": 10, "to_col": 11, "remarks_col": 12},
        "Red Bricks": {"start_row": 3, "date_col": 1, "issue_col": 4, "item_col": 6, "qty_col": 7, "unit_col": 8, "by_col": 9, "to_col": 10, "remarks_col": 11},
        "TMT ": {"start_row": 2, "date_col": 1, "issue_col": 3, "item_col": 5, "spec_col": 6, "qty_col": 7, "unit_col": 8, "by_col": 9, "to_col": 10, "remarks_col": 11},
        "Hume Pipe": {"start_row": 2, "date_col": 1, "issue_col": 2, "item_col": 3, "spec_col": 4, "qty_col": 5, "unit_col": 6, "by_col": 7, "to_col": 8, "remarks_col": 9},
    }
    for sheet_name, config in configs.items():
        import_issue_sheet(store, sheet_name, config)


def seed_budget_heads(store: dict[str, Any]):
    if store.get("budgetHeads"):
        return
    for project in store.get("projects", []):
        store["budgetHeads"].append(
            {
                "id": f"bh-{project['id'].replace('p-', '', 1)}",
                "projectId": project["id"],
                "name": project["name"],
                "amount": project.get("budget", 0),
                "createdBy": "import",
                "createdByName": "Imported Excel record",
                "createdAt": datetime.utcnow().isoformat(),
            }
        )


def main():
    store = load_store()
    ensure_default_users(store)
    reset_imported(store)
    import_oap(store)
    import_issues(store)
    seed_budget_heads(store)
    store["meta"]["sourceFiles"] = [
        "Yarju OAP Inventory Record.xlsx",
        "Yarju Stone, Cement and others record.xlsx",
        "07.docx",
        "Yarju_system.docx",
        "Submission of materials estimate cost for Yarju OAP (1) (3) (1).pdf",
    ]
    store["meta"]["importedAt"] = datetime.utcnow().isoformat()
    STORE.parent.mkdir(parents=True, exist_ok=True)
    STORE.write_text(json.dumps(store, indent=2, default=str), encoding="utf-8")
    print(f"Imported {len(store['items'])} items, {len(store['receipts'])} receipts, {len(store['issues'])} issues, {len(store['requisitions'])} requisitions.")


if __name__ == "__main__":
    main()
